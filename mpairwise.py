import openpyxl
import json

FIRSTWEIGHT = 0.8
SECONDWEIGHT = 2 - FIRSTWEIGHT
OTWIN = 0.55
OTLOSS = 1 - OTWIN
WPWEIGHT = 0.25
OWPWEIGHT = 0.21
OOWPWEIGHT = 0.54
QWBSTART = 0.050
QWBSCALE = QWBSTART / 20

def outputPWR(teamstats, sortedpwr):
    # Create sheet
    wb = openpyxl.Workbook()
    ws = wb.active

    # Header row
    r = 1
    ws.cell(row = r, column = 1).value = 'Rank'
    ws.cell(row = r, column = 2).value = 'Team'
    ws.cell(row = r, column = 3).value = 'PCWs'
    ws.cell(row = r, column = 4).value = 'RPI'
    ws.cell(row = r, column = 5).value = 'QWB'
    ws.cell(row = r, column = 6).value = 'Adj RPI'
    ws.cell(row = r, column = 7).value = 'Unadj RPI'
    # Output each team
    for i in sortedpwr:
        team = i[0]
        pwr, rpi = i[1]
        ws.cell(row = r + 1, column = 1).value = r
        ws.cell(row = r + 1, column = 2).value = team
        ws.cell(row = r + 1, column = 3).value = pwr
        ws.cell(row = r + 1, column = 4).value = rpi
        ws.cell(row = r + 1, column = 5).value = teamstats[team]["qwb"]
        ws.cell(row = r + 1, column = 6).value = teamstats[team]["arpi"]
        ws.cell(row = r + 1, column = 7).value = teamstats[team]["urpi"]
        r += 1

    # Save wb
    wb.save('Womens Pairwise.xlsx')


def compareH2H(teamstats, team1, team2):
    team1wins = countTeam(teamstats[team1]["w"], team2) + countTeam(teamstats[team1]["otw"], team2)
    team2wins = countTeam(teamstats[team2]["w"], team1) + countTeam(teamstats[team2]["otw"], team1)
    return team1wins, team2wins

def compareCoOpp(teamstats, team1, team2):
    # Use set intersect to find common opponents
    coOpp = set(teamstats[team1]["opponents"]).intersection(set(teamstats[team2]["opponents"]))

    # Calculate win percentage against common opponents
    team1CoOpp = 0
    team2CoOpp = 0
    for team in coOpp:
        # Calculate win percentage against common opponent and add to sum
        team1wins = countTeam(teamstats[team1]["w"], team) + countTeam(teamstats[team1]["otw"], team) \
            + 0.5 * countTeam(teamstats[team1]["t"], team)
        team1gp = countTeam(teamstats[team1]["w"], team) + countTeam(teamstats[team1]["l"], team) \
            + countTeam(teamstats[team1]["otw"], team) + countTeam(teamstats[team1]["otl"], team) \
            + countTeam(teamstats[team1]["t"], team)
        team1CoOpp += team1wins / team1gp
        team2wins = countTeam(teamstats[team2]["w"], team) + countTeam(teamstats[team2]["otw"], team) \
            + 0.5 * countTeam(teamstats[team2]["t"], team)
        team2gp = countTeam(teamstats[team2]["w"], team) + countTeam(teamstats[team2]["l"], team) \
            + countTeam(teamstats[team2]["otw"], team) + countTeam(teamstats[team2]["otl"], team) \
            + countTeam(teamstats[team2]["t"], team)
        team2CoOpp += team2wins / team2gp
    
    if team1CoOpp > team2CoOpp:
        return 1, 0
    elif team2CoOpp > team1CoOpp:
        return 0, 1
    else:
        return 0, 0

def compareRPI(teamstats, team1, team2):
    if teamstats[team1]["rpi"] > teamstats[team2]["rpi"]:
        return 1, 0
    else:
        return 0, 1

def calcPWR(teamstats):
    # Stores all comparisons done to save time
    done = set()
    # Compare all teams
    for team1 in teamstats:
        for team2 in teamstats:
            # Check if comparison already made
            if (team1, team2) in done or (team2, team1) in done or team1 == team2:
                continue
            
            # Calculate pairwise components
            rpi1, rpi2 = compareRPI(teamstats, team1, team2)
            coopp1, coopp2 = compareCoOpp(teamstats, team1, team2)
            h2h1, h2h2 = compareH2H(teamstats, team1, team2)
            team1pwr = rpi1 + coopp1 + h2h1
            team2pwr = rpi2 + coopp2 + h2h2

            # Determine winner of comparison
            if team1pwr > team2pwr:
                teamstats[team1]["pwr"] += 1
            elif team2pwr > team1pwr:
                teamstats[team2]["pwr"] += 1
            else:
                if rpi1 > rpi2:
                    teamstats[team1]["pwr"] += 1
                else:
                    teamstats[team2]["pwr"] += 1

            done.add((team1, team2))

def calcFinalRPI(teamstats):
    for key in teamstats:
        teamstats[key]["rpi"] = teamstats[key]["arpi"] + teamstats[key]["qwb"]

def removeBadWins(teamstats):
    for key in teamstats:
        # Initialize adjusted rpi
        teamstats[key]["arpi"] = teamstats[key]["urpi"]
        # Number of changes in an iteration
        changes = 1
        # Sum of rpi removed
        rpisum = 0
        # Number of weighted wins removed
        wgp = 0
        # Set with indices of removed games
        removed = set()
        while changes > 0:
            changes = 0
            # Check each win
            for i, w in enumerate(teamstats[key]["w"]):
                # Check if already removed
                if i in removed:
                    continue
                # Calculate game rpi
                gamerpi, gp = calcGameRPI(teamstats, w, key, "w")
                # Check if game needs to be removed
                if gamerpi + 0.000001 < teamstats[key]["arpi"]:
                    removed.add(i)
                    changes += 1
                    rpisum += gamerpi
                    wgp += gp

            # Update rpi
            teamstats[key]["arpi"] = (teamstats[key]["urpi"] * teamstats[key]["wgp"] - rpisum) / (teamstats[key]["wgp"] - wgp)

def calcQWB(teamstats):
    # Create dict for just rpi
    rpi = dict()
    for t in teamstats:
        rpi[t] = teamstats[t]["arpi"]

    # Sort rpi in descending order to determine qwb for each team
    sortedrpi = sorted(rpi.items(), key=lambda kv: kv[1], reverse = True)
    # Bonus for beating #1 team
    bonus = QWBSTART
    # Key = team; value = bonus for beating that team
    qwb = dict()
    for t in sortedrpi:
        # Make sure bonus isn't negative
        if bonus < 0:
            bonus = 0

        # Add bonus to dict
        qwb[t[0]] = bonus
        # Decrement bonus
        bonus -= QWBSCALE

    # Calculate QWB for each team
    for key in teamstats:
        # Initialize bonus to 0
        teamstats[key]["qwb"] = 0

        # Go through wins and add qwb for each win
        for w in teamstats[key]["w"]:
            if w[1] in qwb:
                teamstats[key]["qwb"] += qwb[w[1]]

        # Go through ot wins and add qwb for each win
        for otw in teamstats[key]["otw"]:
            if otw[1] in qwb:
                teamstats[key]["qwb"] += (qwb[otw[1]] * OTWIN)

        # Go through ot losses and add qwb for each (ot loss counts as OTLOSS wins)
        for otl in teamstats[key]["otl"]:
            if otl[1] in qwb:
                teamstats[key]["qwb"] += (qwb[otl[1]] * OTLOSS)

        # Go through ties and add qwb for each (tie counts as 0.5 wins)
        for t in teamstats[key]["t"]:
            if t[1] in qwb:
                teamstats[key]["qwb"] += (qwb[t[1]] * 0.5)

        # Divide qwb by weighted games played
        teamstats[key]["qwb"] /= teamstats[key]["wgp"]

def calcGameRPI(teamstats, info, team, result):
    if result == "w":
        # Win percentage for that game is 1
        gamerpi = WPWEIGHT + OWPWEIGHT * calcWPwo(teamstats, info[1], team) + OOWPWEIGHT * teamstats[info[1]]["owp"]
        # Calculate rpi weighted for home/road as FIRSTWEIGHT/SECONDWEIGHT
        if info[0] == "H":
            weightedgamerpi = FIRSTWEIGHT * gamerpi
            weightedgp = FIRSTWEIGHT
        elif info[0] == "A":
            weightedgamerpi = SECONDWEIGHT * gamerpi
            weightedgp = SECONDWEIGHT
        else:
            weightedgamerpi = gamerpi
            weightedgp = 1
    elif result == "l":
        # Win percentage for that game is 0
        gamerpi = OWPWEIGHT * calcWPwo(teamstats, info[1], team) + OOWPWEIGHT * teamstats[info[1]]["owp"]
        # Calculate rpi weighted for home/road as SECONDWEIGHT/FIRSTWEIGHT
        if info[0] == "H":
            weightedgamerpi = SECONDWEIGHT * gamerpi
            weightedgp = SECONDWEIGHT
        elif info[0] == "A":
            weightedgamerpi = FIRSTWEIGHT * gamerpi
            weightedgp = FIRSTWEIGHT
        else:
            weightedgamerpi = gamerpi
            weightedgp = 1
    elif result == "otw":
        # Win percentage for that game is OTWIN
        gamerpi = WPWEIGHT * OTWIN + OWPWEIGHT * calcWPwo(teamstats, info[1], team) + OOWPWEIGHT * teamstats[info[1]]["owp"]
        # Calculate rpi weighted for home/road
        if info[0] == "H":
            weightedgamerpi = FIRSTWEIGHT * gamerpi
            weightedgp = FIRSTWEIGHT
        elif info[0] == "A":
            weightedgamerpi = SECONDWEIGHT * gamerpi
            weightedgp = SECONDWEIGHT
        else:
            weightedgamerpi = gamerpi
            weightedgp = 1
        # if info[0] == "H":
        #     weightedgamerpi = (OTWIN * FIRSTWEIGHT + OTLOSS * SECONDWEIGHT) * gamerpi
        #     weightedgp = OTWIN * FIRSTWEIGHT + OTLOSS * SECONDWEIGHT
        # elif info[0] == "A":
        #     weightedgamerpi = (OTWIN * SECONDWEIGHT + OTLOSS * FIRSTWEIGHT) * gamerpi
        #     weightedgp = OTWIN * SECONDWEIGHT + OTLOSS * FIRSTWEIGHT
        # else:
        #     weightedgamerpi = gamerpi
        #     weightedgp = 1
        # weightedgamerpi = gamerpi
        # weightedgp = 1
    elif result == "otl":
        # Win percentage for that game is OTLOSS
        gamerpi = WPWEIGHT * OTLOSS + OWPWEIGHT * calcWPwo(teamstats, info[1], team) + OOWPWEIGHT * teamstats[info[1]]["owp"]
        # Calculate rpi weighted for home/road
        if info[0] == "H":
            weightedgamerpi = SECONDWEIGHT * gamerpi
            weightedgp = SECONDWEIGHT
        elif info[0] == "A":
            weightedgamerpi = FIRSTWEIGHT * gamerpi
            weightedgp = FIRSTWEIGHT
        else:
            weightedgamerpi = gamerpi
            weightedgp = 1
        # if info[0] == "H":
        #     weightedgamerpi = (OTWIN * SECONDWEIGHT + OTLOSS * FIRSTWEIGHT) * gamerpi
        #     weightedgp = OTWIN * SECONDWEIGHT + OTLOSS * FIRSTWEIGHT
        # elif info[0] == "A":
        #     weightedgamerpi = (OTWIN * FIRSTWEIGHT + OTLOSS * SECONDWEIGHT) * gamerpi
        #     weightedgp = OTWIN * FIRSTWEIGHT + OTLOSS * SECONDWEIGHT
        # else:
        #     weightedgamerpi = gamerpi
        #     weightedgp = 1
        # weightedgamerpi = gamerpi
        # weightedgp = 1
    else:
        # Win percentage for that game is 0.5
        gamerpi = WPWEIGHT * 0.5 + OWPWEIGHT * calcWPwo(teamstats, info[1], team) + OOWPWEIGHT * teamstats[info[1]]["owp"]
        weightedgamerpi = gamerpi
        weightedgp = 1
    
    return weightedgamerpi, weightedgp

def calcRPI(teamstats):
    # Calculate components for each team
    calcWP(teamstats)
    calcOWP(teamstats)
    calcOOWP(teamstats)
    
    # Calculate rpi for each team
    for key in teamstats:
        # Calculate rpi for each game
        rpisum = 0
        weightedgp = 0

        # Calculate rpi for each win
        for w in teamstats[key]["w"]:
            gamerpi, gp = calcGameRPI(teamstats, w, key, "w")
            weightedgp += gp
            rpisum += gamerpi

        # Calculate rpi for each loss
        for l in teamstats[key]["l"]:
            gamerpi, gp = calcGameRPI(teamstats, l, key, "l")
            weightedgp += gp
            rpisum += gamerpi

        # Calculate rpi for each ot win
        for otw in teamstats[key]["otw"]:
            gamerpi, gp = calcGameRPI(teamstats, otw, key, "otw")
            weightedgp += gp
            rpisum += gamerpi

        # Calculate rpi for each ot loss
        for otl in teamstats[key]["otl"]:
            gamerpi, gp = calcGameRPI(teamstats, otl, key, "otl")
            weightedgp += gp
            rpisum += gamerpi

        # Calculate rpi for each tie
        for t in teamstats[key]["t"]:
            gamerpi, gp = calcGameRPI(teamstats, t, key, "t")
            weightedgp += gp
            rpisum += gamerpi
        
        teamstats[key]["urpi"] = rpisum / weightedgp
        teamstats[key]["wgp"] = weightedgp

def calcOOWP(teamstats):
    for team in teamstats:
        # Initialize opponent opponent winning percentage
        teamstats[team]["oowp"] = 0
        games = len(teamstats[team]["opponents"])

        # Add opponents
        for opp in teamstats[team]["opponents"]:
            teamstats[team]["oowp"] += teamstats[opp]["owp"]
        
        # Calculate weighted opponent opponent win percentage
        teamstats[team]["oowp"] /= games

def countTeam(teamList, team):
    count = 0
    for tup in teamList:
        if tup[1] == team:
            count += 1
    return count

def calcWPwo(teamstats, team, exclude):
    # Excluding games

    # Initialize winning percentage
    wp = 0
    oppgames = len(teamstats[team]["opponents"]) - teamstats[team]["opponents"].count(exclude)

    # Add wins
    wp += (len(teamstats[team]["w"]) - countTeam(teamstats[team]["w"], exclude))
    
    # Add ot wins: OTWIN of a win and OTLOSS of a loss
    wp += ((len(teamstats[team]["otw"]) - countTeam(teamstats[team]["otw"], exclude)) * OTWIN)
    
    # Add ot losses: OTLOSS of a win and OTWIN of a loss
    wp += ((len(teamstats[team]["otl"]) - countTeam(teamstats[team]["otl"], exclude)) * OTLOSS)
    
    # Add ties: 0.5 of a win and 0.5 of a loss
    wp += ((len(teamstats[team]["t"]) - countTeam(teamstats[team]["t"], exclude)) * 0.5)

    # calculate win percentage
    wp /= oppgames
    return wp

def calcOWP(teamstats):
    for team in teamstats:
        # Initialize opponent winning percentage
        teamstats[team]["owp"] = 0
        games = len(teamstats[team]["opponents"])

        # Add opponents
        for opp in teamstats[team]["opponents"]:
            # Have to manually calculate win percentage with games against current team removed
            wp = calcWPwo(teamstats, opp, team)
            teamstats[team]["owp"] += wp
        
        # calculate opponent win percentage
        teamstats[team]["owp"] /= games

def calcWP(teamstats):
    for team in teamstats:
        # Initialize winning percentage
        teamstats[team]["wp"] = 0
        games = len(teamstats[team]["opponents"])

        # Add wins
        teamstats[team]["wp"] += len(teamstats[team]["w"])
        
        # Add ot wins: OTWIN of a win and OTLOSS of a loss
        teamstats[team]["wp"] += (len(teamstats[team]["otw"]) * OTWIN)
        
        # Add ot losses: OTLOSS of a win and OTWIN of a loss
        teamstats[team]["wp"] += (len(teamstats[team]["otl"]) * OTLOSS)
        
        # Add ties: 0.5 of a win and 0.5 of a loss
        teamstats[team]["wp"] += (len(teamstats[team]["t"]) * 0.5)

        # Calculate win percentage
        teamstats[team]["wp"] /= games

def readGames(teamstats, filename):
    # Open the spreadsheet and assign the first 2 sheets
    # Replace with your xlsx file name
    wb = openpyxl.load_workbook(filename) 
    firstsheet = wb.worksheets[0]

    # Read in every game with the teams
    g = 2
    while g < firstsheet.max_row + 1:
        # Skip exhibition games
        marker = firstsheet.cell(row = g, column = 7).value.strip().lower()
        if marker == 'ex' or marker == 'n3':
            g += 1
            continue

        team1 = firstsheet.cell(row = g, column = 1).value.strip()
        team1score = firstsheet.cell(row = g, column = 2).value
        team2 = firstsheet.cell(row = g, column = 4).value.strip()
        team2score = firstsheet.cell(row = g, column = 5).value
        regulation = firstsheet.cell(row = g, column = 6).value
        # Remove possible padding
        if regulation is not None:
            regulation = regulation.strip()

        # Neutral site game?
        if firstsheet.cell(row = g, column = 3).value.strip() == "vs":
            neutral = True
        else:
            neutral = False
        
        # Initialize if necessary
        if team1 not in teamstats:
            # Initialize team dict
            teamstats[team1] = dict()
            # Initialize comparisons won to 0
            teamstats[team1]["pwr"] = 0
            # Initialize wins, losses, ot wins, ot losses, and ties
            teamstats[team1]["w"] = []
            teamstats[team1]["l"] = []
            teamstats[team1]["otw"] = []
            teamstats[team1]["otl"] = []
            teamstats[team1]["t"] = []
            # Initialize list of opponents
            teamstats[team1]["opponents"] = []
            # Initialize list of future games
            teamstats[team1]["toplay"] = []

        if team2 not in teamstats:
            # Initialize team dict
            teamstats[team2] = dict()
            # Initialize comparisons won to 0
            teamstats[team2]["pwr"] = 0
            # Initialize wins, losses, ot wins, ot losses, and ties
            teamstats[team2]["w"] = []
            teamstats[team2]["l"] = []
            teamstats[team2]["otw"] = []
            teamstats[team2]["otl"] = []
            teamstats[team2]["t"] = []
            # Initialize list of opponents
            teamstats[team2]["opponents"] = []
            # Initialize list of future games
            teamstats[team2]["toplay"] = []

        # Game hasn't been played yet
        if team1score == '' or team1score is None or team2score == '' or team2score is None:
            # update toplay
            if neutral:
                teamstats[team1]["toplay"].append(["N", team2])
                teamstats[team2]["toplay"].append(["N", team1])
            else:
                teamstats[team1]["toplay"].append(["A", team2])
                teamstats[team2]["toplay"].append(["H", team1])
        
        # check to see winner
        if team1score == team2score:
            # update ties
            if neutral:
                teamstats[team1]["t"].append(["N", team2])
                teamstats[team2]["t"].append(["N", team1])
            else:
                teamstats[team1]["t"].append(["A", team2])
                teamstats[team2]["t"].append(["H", team1])
        elif team1score > team2score and (regulation == '' or regulation is None):
            # update wins and losses
            if neutral:
                teamstats[team1]["w"].append(["N", team2])
                teamstats[team2]["l"].append(["N", team1])
            else:
                teamstats[team1]["w"].append(["A", team2])
                teamstats[team2]["l"].append(["H", team1])
        elif team1score < team2score and (regulation == '' or regulation is None):
            # update wins and losses
            if neutral:
                teamstats[team1]["l"].append(["N", team2])
                teamstats[team2]["w"].append(["N", team1])
            else:
                teamstats[team1]["l"].append(["A", team2])
                teamstats[team2]["w"].append(["H", team1])
        elif team1score > team2score:
            # update ot wins and losses
            if neutral:
                teamstats[team1]["otw"].append(["N", team2])
                teamstats[team2]["otl"].append(["N", team1])
            else:
                teamstats[team1]["otw"].append(["A", team2])
                teamstats[team2]["otl"].append(["H", team1])
        else:
            # update ot wins and losses
            if neutral:
                teamstats[team1]["otl"].append(["N", team2])
                teamstats[team2]["otw"].append(["N", team1])
            else:
                teamstats[team1]["otl"].append(["A", team2])
                teamstats[team2]["otw"].append(["H", team1])
        
        # Add opponent
        teamstats[team1]["opponents"].append(team2)
        teamstats[team2]["opponents"].append(team1)
        # Increment counter for reading the spreadsheet
        g += 1

def main():
    # Mega dictionary with all info
    # key = team name
    # value = dictionary with: win pct (float), opponent win pct (float), opponent opponent win pct (float)
    # quality wins bonus (float), wins (list of opponents beaten), losses (list of opponents lost to),
    # overtime wins (list of opponents beaten in ot), pairwise comparisons won (int),
    # opponents (list of opponents), weighted games played (float)
    teamstats = dict()

    readGames(teamstats, "../NCAA games.xlsx")
    calcRPI(teamstats)
    removeBadWins(teamstats)
    calcQWB(teamstats)
    calcFinalRPI(teamstats)
    calcPWR(teamstats)

    # For output in order of the pairwise
    pwr = dict()
    for t in teamstats:
        pwr[t] = (teamstats[t]["pwr"], teamstats[t]["rpi"])
    sortedpwr = sorted(pwr.items(), key=lambda kv: (kv[1][0], kv[1][1]), reverse = True)
    outputPWR(teamstats, sortedpwr)
    return sortedpwr

if __name__ == "__main__":
    main()
